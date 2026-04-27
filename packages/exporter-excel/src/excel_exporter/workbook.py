"""Generate the catalog workbook from a snapshot."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import openpyxl

from excel_exporter.models import CatalogSnapshot
from excel_exporter.tabs import (
    write_actors,
    write_assets,
    write_lineage,
    write_overview,
    write_relations,
    write_schemas,
)


def _build(snapshot: CatalogSnapshot) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    # First sheet exists by default; we rename + use it for Overview, then add the rest.
    write_overview(wb.active, snapshot)
    write_assets(wb.create_sheet(), snapshot)
    write_actors(wb.create_sheet(), snapshot)
    write_relations(wb.create_sheet(), snapshot)
    write_schemas(wb.create_sheet(), snapshot)
    write_lineage(wb.create_sheet(), snapshot)
    return wb


def write_workbook(snapshot: CatalogSnapshot, output_path: str | Path) -> None:
    """Write the snapshot to a .xlsx file at the given path."""
    wb = _build(snapshot)
    wb.save(str(output_path))


def write_workbook_to_bytes(snapshot: CatalogSnapshot) -> bytes:
    """Return the workbook as bytes (for streaming responses)."""
    wb = _build(snapshot)
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
