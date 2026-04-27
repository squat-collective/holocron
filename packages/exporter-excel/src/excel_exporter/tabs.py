"""One function per workbook tab — pure, take a snapshot, write to a worksheet.

Verbosity rule: every UID surfaced in the workbook is paired with a human-
readable label column next to it. Naked UIDs are unhelpful — the label
turns "act1, aaa1, owns" into "Jean Dupont (person) owns customers.xlsx
(dataset)" at a glance.
"""

from __future__ import annotations

import json
from typing import Any

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from excel_exporter.models import (
    ActorRecord,
    AssetRecord,
    CatalogSnapshot,
)

_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

# Type emojis to make tabs scannable at a glance
_ASSET_TYPE_GLYPH = {
    "dataset": "📊",
    "report": "📈",
    "process": "⚙️",
    "system": "🖥️",
}
_ACTOR_TYPE_GLYPH = {"person": "👤", "group": "👥"}
_RELATION_TYPE_GLYPH = {
    "owns": "🔑",
    "uses": "👀",
    "feeds": "➡️",
    "contains": "📦",
    "member_of": "🧑\u200d🤝\u200d🧑",
    "applies_to": "🛡️",
}


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def _autosize(ws: Worksheet, max_width: int = 60) -> None:
    """Cheap auto-size: sample first 50 rows + header per column."""
    if ws.max_row is None or ws.max_column is None:
        return
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, min(ws.max_row, 50) + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            length = len(str(v))
            if length > max_len:
                max_len = length
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, max_width)


def _verified_glyph(verified: bool) -> str:
    return "✅" if verified else "⚠️"


def _entity_label(uid: str, asset_lookup: dict[str, AssetRecord], actor_lookup: dict[str, ActorRecord]) -> str:
    """Resolve a UID into a verbose label like '📊 customers.xlsx (dataset)'.

    Tries assets first, then actors, then falls back to the bare UID.
    Used everywhere a UID appears in a relation/cross-reference.
    """
    asset = asset_lookup.get(uid)
    if asset is not None:
        glyph = _ASSET_TYPE_GLYPH.get(asset.type, "📦")
        return f"{glyph} {asset.name} ({asset.type})"
    actor = actor_lookup.get(uid)
    if actor is not None:
        glyph = _ACTOR_TYPE_GLYPH.get(actor.type, "👤")
        email = f" <{actor.email}>" if actor.email else ""
        return f"{glyph} {actor.name}{email} ({actor.type})"
    return f"❓ <unknown:{uid[:12]}>"


def _asset_summary(a: AssetRecord) -> str:
    """One-line human description of what an asset is — derived from metadata.

    For excel-discovered assets this surfaces sheet/table/field counts so you
    can scan the Assets tab without opening the JSON metadata column.
    """
    bits: list[str] = []
    schema = a.metadata.get("schema") if isinstance(a.metadata, dict) else None
    if isinstance(schema, list) and schema:
        sheets = sum(1 for n in schema if isinstance(n, dict) and n.get("nodeType") == "container")
        tables = 0
        fields = 0
        stack: list[Any] = list(schema)
        while stack:
            node = stack.pop()
            if not isinstance(node, dict):
                continue
            if node.get("nodeType") == "field":
                fields += 1
            elif node.get("containerType") == "table":
                tables += 1
            children = node.get("children")
            if isinstance(children, list):
                stack.extend(children)
        bits.append(f"{sheets} sheet{'s' if sheets != 1 else ''}")
        if tables:
            bits.append(f"{tables} table{'s' if tables != 1 else ''}")
        if fields:
            bits.append(f"{fields} field{'s' if fields != 1 else ''}")
    if isinstance(a.metadata, dict):
        hints = a.metadata.get("lineage_hints")
        if isinstance(hints, list) and hints:
            bits.append(f"{len(hints)} lineage hint{'s' if len(hints) != 1 else ''}")
        ext_links = a.metadata.get("external_links")
        if isinstance(ext_links, list) and ext_links:
            bits.append(f"{len(ext_links)} external ref{'s' if len(ext_links) != 1 else ''}")
        if a.metadata.get("discovered_via") == "external_link":
            bits.append("(external workbook ref)")
    return " · ".join(bits)


# ----- Overview -----

def write_overview(ws: Worksheet, snapshot: CatalogSnapshot) -> None:
    ws.title = "Overview"

    # Break down counts by entity sub-type so readers see what kind of data is in here
    asset_types = sorted({a.type for a in snapshot.assets})
    actor_types = sorted({a.type for a in snapshot.actors})
    rel_types = sorted({r.type for r in snapshot.relations})

    rows: list[tuple[str, Any]] = [
        ("Generated at", snapshot.fetched_at.isoformat(timespec="seconds")),
        ("Source API", snapshot.api_url),
        ("", ""),
        ("Assets — total", len(snapshot.assets)),
        ("Assets — unverified", sum(1 for a in snapshot.assets if not a.verified)),
    ]
    for t in asset_types:
        glyph = _ASSET_TYPE_GLYPH.get(t, "📦")
        rows.append(
            (f"  {glyph} {t}", sum(1 for a in snapshot.assets if a.type == t))
        )

    rows += [
        ("", ""),
        ("Actors — total", len(snapshot.actors)),
        ("Actors — unverified", sum(1 for a in snapshot.actors if not a.verified)),
    ]
    for t in actor_types:
        glyph = _ACTOR_TYPE_GLYPH.get(t, "👤")
        rows.append(
            (f"  {glyph} {t}", sum(1 for a in snapshot.actors if a.type == t))
        )

    rows += [
        ("", ""),
        ("Relations — total", len(snapshot.relations)),
        ("Relations — unverified", sum(1 for r in snapshot.relations if not r.verified)),
    ]
    for t in rel_types:
        glyph = _RELATION_TYPE_GLYPH.get(t, "🔗")
        rows.append(
            (f"  {glyph} {t}", sum(1 for r in snapshot.relations if r.type == t))
        )

    # Summary insights derived from metadata
    discovered_by_counts: dict[str, int] = {}
    for a in snapshot.assets:
        key = a.discovered_by or "(manual)"
        discovered_by_counts[key] = discovered_by_counts.get(key, 0) + 1
    if discovered_by_counts:
        rows += [("", ""), ("Discovered by", "")]
        for source, n in sorted(discovered_by_counts.items()):
            rows.append((f"  {source}", n))

    ws.cell(row=1, column=1, value="Holocron catalog export").font = Font(size=16, bold=True)
    for r, (label, value) in enumerate(rows, start=3):
        cell = ws.cell(row=r, column=1, value=label)
        if not label.startswith("  "):
            cell.font = Font(bold=True)
        ws.cell(row=r, column=2, value=value)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60


# ----- Assets -----

def write_assets(ws: Worksheet, snapshot: CatalogSnapshot) -> None:
    ws.title = "Assets"
    headers = [
        "Status",
        "label",  # verbose: "📊 customers.xlsx (dataset)" — at-a-glance identity
        "uid",
        "type",
        "name",
        "summary",  # derived: "2 sheets · 2 tables · 6 fields · 1 lineage hint"
        "description",
        "location",
        "lifecycle",
        "verified",
        "discovered_by",
        "created_at",
        "updated_at",
        "metadata (JSON)",
    ]
    _write_header(ws, headers)
    for r, a in enumerate(snapshot.assets, start=2):
        glyph = _ASSET_TYPE_GLYPH.get(a.type, "📦")
        ws.cell(row=r, column=1, value=_verified_glyph(a.verified))
        ws.cell(row=r, column=2, value=f"{glyph} {a.name} ({a.type})")
        ws.cell(row=r, column=3, value=a.uid)
        ws.cell(row=r, column=4, value=a.type)
        ws.cell(row=r, column=5, value=a.name)
        ws.cell(row=r, column=6, value=_asset_summary(a))
        ws.cell(row=r, column=7, value=a.description)
        ws.cell(row=r, column=8, value=a.location)
        ws.cell(row=r, column=9, value=a.status)
        ws.cell(row=r, column=10, value=a.verified)
        ws.cell(row=r, column=11, value=a.discovered_by)
        ws.cell(row=r, column=12, value=a.created_at.isoformat(timespec="seconds"))
        ws.cell(row=r, column=13, value=a.updated_at.isoformat(timespec="seconds"))
        ws.cell(row=r, column=14, value=json.dumps(a.metadata, ensure_ascii=False, default=str))
    ws.freeze_panes = "A2"
    _autosize(ws)


# ----- Actors -----

def write_actors(ws: Worksheet, snapshot: CatalogSnapshot) -> None:
    ws.title = "Actors"

    # Pre-count relations per actor uid → so we can show "owns 3, uses 1"
    relations_by_actor: dict[str, dict[str, int]] = {}
    for rel in snapshot.relations:
        bucket = relations_by_actor.setdefault(rel.from_uid, {})
        bucket[rel.type] = bucket.get(rel.type, 0) + 1

    headers = [
        "Status",
        "label",  # verbose: "👤 Jean Dupont <jean@acme.com> (person)"
        "uid",
        "type",
        "name",
        "email",
        "activity",  # derived: "owns 3, uses 1"
        "description",
        "verified",
        "discovered_by",
        "created_at",
        "metadata (JSON)",
    ]
    _write_header(ws, headers)
    for r, a in enumerate(snapshot.actors, start=2):
        glyph = _ACTOR_TYPE_GLYPH.get(a.type, "👤")
        email_part = f" <{a.email}>" if a.email else ""
        activity = ", ".join(
            f"{rt} {n}" for rt, n in sorted(relations_by_actor.get(a.uid, {}).items())
        )
        ws.cell(row=r, column=1, value=_verified_glyph(a.verified))
        ws.cell(row=r, column=2, value=f"{glyph} {a.name}{email_part} ({a.type})")
        ws.cell(row=r, column=3, value=a.uid)
        ws.cell(row=r, column=4, value=a.type)
        ws.cell(row=r, column=5, value=a.name)
        ws.cell(row=r, column=6, value=a.email)
        ws.cell(row=r, column=7, value=activity)
        ws.cell(row=r, column=8, value=a.description)
        ws.cell(row=r, column=9, value=a.verified)
        ws.cell(row=r, column=10, value=a.discovered_by)
        ws.cell(row=r, column=11, value=a.created_at.isoformat(timespec="seconds"))
        ws.cell(row=r, column=12, value=json.dumps(a.metadata, ensure_ascii=False, default=str))
    ws.freeze_panes = "A2"
    _autosize(ws)


# ----- Relations -----

def write_relations(ws: Worksheet, snapshot: CatalogSnapshot) -> None:
    ws.title = "Relations"

    asset_lookup = {a.uid: a for a in snapshot.assets}
    actor_lookup = {a.uid: a for a in snapshot.actors}

    headers = [
        "Status",
        "sentence",  # verbose: "👤 Jean Dupont (person) ─ owns ─▶ 📊 customers.xlsx (dataset)"
        "type",
        "from",
        "to",
        "uid",
        "from_uid",
        "to_uid",
        "verified",
        "discovered_by",
        "created_at",
        "properties (JSON)",
    ]
    _write_header(ws, headers)
    for r, rel in enumerate(snapshot.relations, start=2):
        from_label = _entity_label(rel.from_uid, asset_lookup, actor_lookup)
        to_label = _entity_label(rel.to_uid, asset_lookup, actor_lookup)
        glyph = _RELATION_TYPE_GLYPH.get(rel.type, "🔗")
        sentence = f"{from_label}  ─ {glyph} {rel.type} ─▶  {to_label}"

        ws.cell(row=r, column=1, value=_verified_glyph(rel.verified))
        ws.cell(row=r, column=2, value=sentence)
        ws.cell(row=r, column=3, value=rel.type)
        ws.cell(row=r, column=4, value=from_label)
        ws.cell(row=r, column=5, value=to_label)
        ws.cell(row=r, column=6, value=rel.uid)
        ws.cell(row=r, column=7, value=rel.from_uid)
        ws.cell(row=r, column=8, value=rel.to_uid)
        ws.cell(row=r, column=9, value=rel.verified)
        ws.cell(row=r, column=10, value=rel.discovered_by)
        ws.cell(row=r, column=11, value=rel.created_at.isoformat(timespec="seconds"))
        ws.cell(row=r, column=12, value=json.dumps(rel.properties, ensure_ascii=False, default=str))
    ws.freeze_panes = "A2"
    _autosize(ws, max_width=80)


# ----- Schemas (flattened) -----

def _flatten_fields(
    nodes: list[dict[str, Any]] | None, path: list[str] | None = None
) -> list[dict[str, Any]]:
    """Walk a SchemaNode tree, return one row per leaf field."""
    out: list[dict[str, Any]] = []
    if not nodes:
        return out
    path = path or []
    for node in nodes:
        if not isinstance(node, dict):
            continue
        name = str(node.get("name") or "")
        nt = node.get("nodeType")
        children = node.get("children")
        if nt == "field":
            out.append(
                {
                    "path": "/".join([*path, name]),
                    "name": name,
                    "dataType": node.get("dataType") or "",
                    "pii": bool(node.get("pii", False)),
                    "description": node.get("description") or "",
                }
            )
        if isinstance(children, list):
            out.extend(_flatten_fields(children, [*path, name]))
    return out


def write_schemas(ws: Worksheet, snapshot: CatalogSnapshot) -> None:
    ws.title = "Schemas"
    headers = ["Asset", "Asset uid", "Field path", "Field name", "Data type", "PII", "Description"]
    _write_header(ws, headers)

    row = 2
    for asset in snapshot.assets:
        schema = asset.metadata.get("schema") if isinstance(asset.metadata, dict) else None
        if not isinstance(schema, list):
            continue
        for f in _flatten_fields(schema):
            ws.cell(row=row, column=1, value=asset.name)
            ws.cell(row=row, column=2, value=asset.uid)
            ws.cell(row=row, column=3, value=f["path"])
            ws.cell(row=row, column=4, value=f["name"])
            ws.cell(row=row, column=5, value=f["dataType"])
            ws.cell(row=row, column=6, value="✓" if f["pii"] else "")
            ws.cell(row=row, column=7, value=f["description"])
            row += 1
    ws.freeze_panes = "A2"
    _autosize(ws)


# ----- Lineage -----

def write_lineage(ws: Worksheet, snapshot: CatalogSnapshot) -> None:
    ws.title = "Lineage"
    headers = [
        "sentence",  # verbose narrative line
        "source",  # labeled
        "target",  # labeled
        "kind",
        "via",
        "source uid",
        "target uid",
    ]
    _write_header(ws, headers)

    asset_lookup = {a.uid: a for a in snapshot.assets}
    actor_lookup = {a.uid: a for a in snapshot.actors}
    row = 2

    # 1. 'feeds' relations from the graph
    for rel in snapshot.relations:
        if rel.type != "feeds":
            continue
        src = _entity_label(rel.from_uid, asset_lookup, actor_lookup)
        tgt = _entity_label(rel.to_uid, asset_lookup, actor_lookup)
        via = str(rel.properties.get("via_formula", "")) or "(no formula recorded)"
        ws.cell(row=row, column=1, value=f"{src}  ─ ➡️ feeds ─▶  {tgt}")
        ws.cell(row=row, column=2, value=src)
        ws.cell(row=row, column=3, value=tgt)
        ws.cell(row=row, column=4, value="cross-asset feed")
        ws.cell(row=row, column=5, value=via)
        ws.cell(row=row, column=6, value=rel.from_uid)
        ws.cell(row=row, column=7, value=rel.to_uid)
        row += 1

    # 2. In-asset lineage_hints from metadata (e.g. cross-sheet VLOOKUP within one workbook)
    for asset in snapshot.assets:
        hints = asset.metadata.get("lineage_hints") if isinstance(asset.metadata, dict) else None
        if not isinstance(hints, list):
            continue
        for h in hints:
            if not isinstance(h, dict):
                continue
            from_label = f"📊 {asset.name} → {h.get('from_sheet', '')}"
            to_label = (
                f"📊 {asset.name} → {h.get('to_sheet', '')} → {h.get('to_table', '')}"
                f"!{h.get('to_cell', '')}"
            )
            via = h.get("via_formula") or ""
            kind = "lookup hint" if h.get("is_lookup") else "in-asset hint"
            ws.cell(row=row, column=1, value=f"{from_label}  ─ ➡️ {kind} ─▶  {to_label}")
            ws.cell(row=row, column=2, value=from_label)
            ws.cell(row=row, column=3, value=to_label)
            ws.cell(row=row, column=4, value=kind)
            ws.cell(row=row, column=5, value=via)
            ws.cell(row=row, column=6, value=asset.uid)
            ws.cell(row=row, column=7, value=asset.uid)
            row += 1

    ws.freeze_panes = "A2"
    _autosize(ws, max_width=80)
