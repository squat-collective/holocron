"""End-to-end test: build a workbook from a snapshot, then read it back."""

from pathlib import Path

import openpyxl

from excel_exporter import write_workbook
from excel_exporter.models import CatalogSnapshot


def _open(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=True)


def test_all_six_tabs_present(snapshot: CatalogSnapshot, tmp_path: Path) -> None:
    out = tmp_path / "catalog.xlsx"
    write_workbook(snapshot, out)
    wb = _open(out)
    assert wb.sheetnames == ["Overview", "Assets", "Actors", "Relations", "Schemas", "Lineage"]


def test_overview_has_counts(snapshot: CatalogSnapshot, tmp_path: Path) -> None:
    out = tmp_path / "catalog.xlsx"
    write_workbook(snapshot, out)
    ws = _open(out)["Overview"]
    # Walk all label/value rows and look up by label (layout has variable size now
    # because of per-type breakdowns).
    flat = {
        ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
        for r in range(3, ws.max_row + 1)
    }
    assert flat["Assets — total"] == 2
    assert flat["Assets — unverified"] == 1
    assert flat["Actors — total"] == 1
    assert flat["Relations — total"] == 2


def test_assets_tab_has_verbose_label_and_summary(
    snapshot: CatalogSnapshot, tmp_path: Path
) -> None:
    out = tmp_path / "catalog.xlsx"
    write_workbook(snapshot, out)
    ws = _open(out)["Assets"]
    # Header + 2 data rows
    assert ws.max_row == 3
    # Status column reflects verified flag
    assert ws.cell(row=2, column=1).value == "⚠️"
    assert ws.cell(row=3, column=1).value == "✅"

    # Label column resolves uid → "📊 customers.xlsx (dataset)"
    label_first = ws.cell(row=2, column=2).value
    assert label_first is not None
    assert "customers.xlsx" in label_first
    assert "dataset" in label_first
    assert "📊" in label_first

    # Summary column derives from metadata.schema → field counts visible at a glance
    summary_first = ws.cell(row=2, column=6).value
    assert summary_first is not None
    assert "1 sheet" in summary_first
    assert "1 table" in summary_first
    assert "2 fields" in summary_first
    assert "1 lineage hint" in summary_first


def test_relations_tab_resolves_uids_to_labels(
    snapshot: CatalogSnapshot, tmp_path: Path
) -> None:
    out = tmp_path / "catalog.xlsx"
    write_workbook(snapshot, out)
    ws = _open(out)["Relations"]

    # Header + 2 data rows
    assert ws.max_row == 3

    # Find the 'owns' relation row
    owns_row = next(
        r for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=3).value == "owns"
    )
    sentence = ws.cell(row=owns_row, column=2).value
    from_label = ws.cell(row=owns_row, column=4).value
    to_label = ws.cell(row=owns_row, column=5).value

    # The actor's name + the asset's name should appear in BOTH the sentence and
    # the labeled from/to columns
    assert "Jean Dupont" in (from_label or "")
    assert "customers.xlsx" in (to_label or "")
    assert "Jean Dupont" in (sentence or "")
    assert "owns" in (sentence or "")
    assert "customers.xlsx" in (sentence or "")


def test_actors_tab_shows_activity_summary(
    snapshot: CatalogSnapshot, tmp_path: Path
) -> None:
    out = tmp_path / "catalog.xlsx"
    write_workbook(snapshot, out)
    ws = _open(out)["Actors"]
    # Activity column should say "owns 1" because the actor in the fixture owns one asset
    activity = ws.cell(row=2, column=7).value
    assert activity == "owns 1"
    # Label column has emoji + name + email + type
    label = ws.cell(row=2, column=2).value
    assert label is not None
    assert "Jean Dupont" in label
    assert "jean@acme.com" in label
    assert "person" in label


def test_schemas_tab_flattens_fields(snapshot: CatalogSnapshot, tmp_path: Path) -> None:
    out = tmp_path / "catalog.xlsx"
    write_workbook(snapshot, out)
    ws = _open(out)["Schemas"]
    rows = [
        {
            "asset": ws.cell(row=r, column=1).value,
            "path": ws.cell(row=r, column=3).value,
            "name": ws.cell(row=r, column=4).value,
            "dtype": ws.cell(row=r, column=5).value,
            "pii": ws.cell(row=r, column=6).value,
        }
        for r in range(2, ws.max_row + 1)
    ]
    # Two fields from the customers schema
    assert len(rows) == 2
    paths = {r["path"] for r in rows}
    assert paths == {"Customers/Customers/id", "Customers/Customers/email"}
    pii_row = next(r for r in rows if r["name"] == "email")
    assert pii_row["pii"] == "✓"
    assert pii_row["dtype"] == "string"


def test_lineage_tab_unions_feeds_and_hints(snapshot: CatalogSnapshot, tmp_path: Path) -> None:
    out = tmp_path / "catalog.xlsx"
    write_workbook(snapshot, out)
    ws = _open(out)["Lineage"]
    # Layout: sentence | source | target | kind | via | source_uid | target_uid
    rows = [
        {
            "sentence": ws.cell(row=r, column=1).value,
            "source": ws.cell(row=r, column=2).value,
            "target": ws.cell(row=r, column=3).value,
            "kind": ws.cell(row=r, column=4).value,
            "via": ws.cell(row=r, column=5).value,
        }
        for r in range(2, ws.max_row + 1)
    ]
    kinds = {r["kind"] for r in rows}
    assert "cross-asset feed" in kinds  # from feeds relation
    assert "lookup hint" in kinds  # the fixture's lineage_hint has is_lookup=True
    # Sentence column resolves uids → labels with emoji
    sentences = [r["sentence"] for r in rows if r["sentence"]]
    assert any("customers.xlsx" in s and "manual" in s for s in sentences)


def test_relations_tab_counts(snapshot: CatalogSnapshot, tmp_path: Path) -> None:
    out = tmp_path / "catalog.xlsx"
    write_workbook(snapshot, out)
    ws = _open(out)["Relations"]
    assert ws.max_row == 3  # header + 2 relations
