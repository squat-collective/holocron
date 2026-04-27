"""Unit tests for the compliance analyzers + workbook smoke check."""

from __future__ import annotations

import io
import zipfile
from datetime import UTC, datetime

from compliance_report.analysis import (
    build_coverage,
    build_ownership,
    build_pii_inventory,
    build_rule_applications,
    build_verifications,
    run_report,
)
from compliance_report.models import (
    ActorRecord,
    AssetRecord,
    CatalogSnapshot,
    RelationRecord,
    RuleRecord,
)
from compliance_report.workbook import write_compliance_xlsx

NOW = datetime(2026, 4, 25, 12, 0, 0, tzinfo=UTC)


def _asset(uid: str, **over: object) -> AssetRecord:
    base = {
        "uid": uid,
        "type": "dataset",
        "name": uid,
        "status": "active",
        "verified": True,
        "created_at": NOW,
        "updated_at": NOW,
    }
    base.update(over)  # type: ignore[arg-type]
    return AssetRecord.model_validate(base)


def _actor(uid: str, **over: object) -> ActorRecord:
    base = {
        "uid": uid,
        "type": "person",
        "name": uid,
        "verified": True,
        "created_at": NOW,
        "updated_at": NOW,
    }
    base.update(over)  # type: ignore[arg-type]
    return ActorRecord.model_validate(base)


def _rule(uid: str, **over: object) -> RuleRecord:
    base = {
        "uid": uid,
        "name": uid,
        "severity": "warning",
        "verified": True,
        "created_at": NOW,
        "updated_at": NOW,
    }
    base.update(over)  # type: ignore[arg-type]
    return RuleRecord.model_validate(base)


def _rel(from_uid: str, to_uid: str, type_: str, **over: object) -> RelationRecord:
    base = {
        "uid": f"rel-{from_uid}-{type_}-{to_uid}",
        "from_uid": from_uid,
        "to_uid": to_uid,
        "type": type_,
        "verified": True,
        "created_at": NOW,
        "properties": over.get("properties", {}),
    }
    return RelationRecord.model_validate(base)


# ---------- coverage ----------


class TestBuildCoverage:
    def test_empty_catalog_yields_zero_pcts(self) -> None:
        c = build_coverage(CatalogSnapshot(fetched_at=NOW))
        assert c.assets_total == 0
        # Empty denominator → 0.0, never NaN. Asserted because Excel
        # users sort/filter on this column and NaN breaks both.
        assert c.pct(c.assets_with_owner, c.assets_total) == 0.0

    def test_owner_coverage_count(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a"), _asset("b"), _asset("c")],
            actors=[_actor("p")],
            relations=[_rel("p", "a", "owns")],
        )
        c = build_coverage(snap)
        assert c.assets_with_owner == 1
        assert c.assets_total == 3

    def test_description_coverage_treats_empty_string_as_missing(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[
                _asset("a", description="real"),
                _asset("b", description=""),
                _asset("c"),
            ],
        )
        c = build_coverage(snap)
        assert c.assets_with_description == 1

    def test_pii_field_count_walks_nested_schema(self) -> None:
        schema = [
            {
                "name": "Customers",
                "nodeType": "container",
                "children": [
                    {"name": "email", "nodeType": "field", "pii": True},
                    {"name": "country", "nodeType": "field"},
                    {
                        "name": "address",
                        "nodeType": "container",
                        "children": [
                            {"name": "zip", "nodeType": "field", "pii": True},
                        ],
                    },
                ],
            }
        ]
        snap = CatalogSnapshot(
            fetched_at=NOW, assets=[_asset("a", metadata={"schema": schema})]
        )
        c = build_coverage(snap)
        # Two fields flagged pii=True, one nested inside a sub-container.
        assert c.pii_field_count == 2

    def test_rules_applied_counts_distinct_rules(self) -> None:
        # Two `applies_to` for the same rule should still count it once.
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a"), _asset("b")],
            rules=[_rule("r1"), _rule("r2")],
            relations=[
                _rel("r1", "a", "applies_to"),
                _rel("r1", "b", "applies_to"),
            ],
        )
        c = build_coverage(snap)
        assert c.rules_applied == 1


# ---------- rule applications ----------


class TestBuildRuleApplications:
    def test_picks_up_enforcement_from_relation_properties(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a", name="Sales")],
            rules=[_rule("r", name="Has owner", severity="critical", category="ownership")],
            relations=[
                _rel(
                    "r",
                    "a",
                    "applies_to",
                    properties={"enforcement": "enforced", "field_path": "email"},
                )
            ],
        )
        out = build_rule_applications(snap)
        assert len(out) == 1
        row = out[0]
        assert row.enforcement == "enforced"
        assert row.field_path == "email"
        assert row.category == "ownership"

    def test_sorts_critical_first_then_alpha(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a")],
            rules=[
                _rule("r-info", name="Z info", severity="info"),
                _rule("r-crit", name="A critical", severity="critical"),
                _rule("r-warn", name="M warning", severity="warning"),
            ],
            relations=[
                _rel("r-info", "a", "applies_to"),
                _rel("r-crit", "a", "applies_to"),
                _rel("r-warn", "a", "applies_to"),
            ],
        )
        out = build_rule_applications(snap)
        assert [r.severity for r in out] == ["critical", "warning", "info"]

    def test_skips_unknown_uids(self) -> None:
        # An applies_to relation pointing at a deleted rule shouldn't crash
        # the report; just drop the row.
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a")],
            rules=[],
            relations=[_rel("r-unknown", "a", "applies_to")],
        )
        assert build_rule_applications(snap) == []


# ---------- PII inventory ----------


class TestBuildPiiInventory:
    def test_emits_one_row_per_pii_field_with_path(self) -> None:
        schema = [
            {
                "name": "Users",
                "nodeType": "container",
                "children": [
                    {"name": "email", "nodeType": "field", "dataType": "string", "pii": True},
                    {"name": "id", "nodeType": "field"},
                ],
            }
        ]
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a", name="People DB", metadata={"schema": schema})],
        )
        out = build_pii_inventory(snap)
        assert len(out) == 1
        row = out[0]
        assert row.asset_name == "People DB"
        assert row.field_path == "Users/email"
        assert row.data_type == "string"

    def test_joins_owner_names_when_available(self) -> None:
        schema = [{"name": "email", "nodeType": "field", "pii": True}]
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a", metadata={"schema": schema})],
            actors=[_actor("p1", name="Alice"), _actor("p2", name="Bob")],
            relations=[_rel("p1", "a", "owns"), _rel("p2", "a", "owns")],
        )
        out = build_pii_inventory(snap)
        # Owner names sorted by relation insertion order (which mirrors
        # Cypher result order in the live API). Either ordering acceptable
        # — what matters is both names appear.
        assert "Alice" in out[0].owner_names
        assert "Bob" in out[0].owner_names

    def test_skips_assets_without_pii(self) -> None:
        snap = CatalogSnapshot(fetched_at=NOW, assets=[_asset("a")])
        assert build_pii_inventory(snap) == []


# ---------- ownership ----------


class TestBuildOwnership:
    def test_flattens_owns_edges_with_names(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a1", name="Sales")],
            actors=[_actor("p1", name="Alice")],
            relations=[_rel("p1", "a1", "owns")],
        )
        out = build_ownership(snap)
        assert len(out) == 1
        assert out[0].actor_name == "Alice"
        assert out[0].asset_name == "Sales"


# ---------- verifications ----------


class TestBuildVerifications:
    def test_only_verified_entities_appear(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a", verified=True), _asset("b", verified=False)],
            actors=[_actor("p", verified=True)],
        )
        out = build_verifications(snap)
        names = sorted(v.entity_name for v in out)
        assert names == ["a", "p"]

    def test_sorted_newest_first(self) -> None:
        older = datetime(2026, 1, 1, tzinfo=UTC)
        newer = datetime(2026, 4, 1, tzinfo=UTC)
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[
                _asset("old", updated_at=older),
                _asset("new", updated_at=newer),
            ],
        )
        out = build_verifications(snap)
        assert [v.entity_name for v in out] == ["new", "old"]


# ---------- run_report + workbook ----------


class TestRunReport:
    def test_bundles_every_section(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a", description="ok", metadata={"schema": [{"name": "f", "nodeType": "field", "pii": True}]})],
            actors=[_actor("p")],
            rules=[_rule("r")],
            relations=[_rel("p", "a", "owns"), _rel("r", "a", "applies_to")],
        )
        report = run_report(snap)
        assert report.coverage.assets_with_owner == 1
        assert len(report.rule_applications) == 1
        assert len(report.pii_fields) == 1
        assert len(report.ownership_rows) == 1
        assert len(report.verifications) == 3  # asset + actor + rule


class TestWriteComplianceXlsx:
    def test_produces_a_valid_xlsx_with_all_sheets(self) -> None:
        body = write_compliance_xlsx(
            run_report(CatalogSnapshot(fetched_at=NOW)), generated_at=NOW
        )
        assert body[:2] == b"PK"
        with zipfile.ZipFile(io.BytesIO(body)) as zf:
            names = zf.namelist()
        assert any(n.endswith("workbook.xml") for n in names)

        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(body))
        assert wb.sheetnames == [
            "Overview",
            "Rules in force",
            "PII inventory",
            "Ownership",
            "Verifications",
        ]
