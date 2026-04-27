"""Render a `ComplianceReport` as an xlsx workbook.

Layout:
    Overview            — coverage % + counts
    Rules in force      — every applied rule with target asset + enforcement
    PII inventory       — every PII-flagged schema field, with owners
    Ownership           — flat actor → asset matrix
    Verifications       — verified entities sorted newest-first
"""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from compliance_report.analysis import (
    ComplianceReport,
    CoverageStats,
    OwnershipRow,
    PIIField,
    RuleApplication,
    VerificationEntry,
)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # slate-800


def write_compliance_xlsx(report: ComplianceReport, *, generated_at: datetime) -> bytes:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_overview(wb.create_sheet("Overview"), report.coverage, generated_at)
    _write_rules(wb.create_sheet("Rules in force"), report.rule_applications)
    _write_pii(wb.create_sheet("PII inventory"), report.pii_fields)
    _write_ownership(wb.create_sheet("Ownership"), report.ownership_rows)
    _write_verifications(wb.create_sheet("Verifications"), report.verifications)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------- Overview ----------


def _write_overview(ws: Worksheet, c: CoverageStats, generated_at: datetime) -> None:
    ws.append(["Holocron compliance report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Generated at {generated_at.isoformat()}"])
    ws.append([])

    ws.append(["Metric", "Numerator", "Denominator", "%"])
    _stripe_header(ws, row=4, cols=4)

    ws.append(
        ["Assets with an owner", c.assets_with_owner, c.assets_total, c.pct(c.assets_with_owner, c.assets_total)]
    )
    ws.append(
        [
            "Assets with a description",
            c.assets_with_description,
            c.assets_total,
            c.pct(c.assets_with_description, c.assets_total),
        ]
    )
    ws.append(
        ["Assets verified", c.assets_verified, c.assets_total, c.pct(c.assets_verified, c.assets_total)]
    )
    ws.append(
        ["Actors verified", c.actors_verified, c.actors_total, c.pct(c.actors_verified, c.actors_total)]
    )
    ws.append(
        ["Rules verified", c.rules_verified, c.rules_total, c.pct(c.rules_verified, c.rules_total)]
    )
    ws.append(
        ["Rules applied", c.rules_applied, c.rules_total, c.pct(c.rules_applied, c.rules_total)]
    )

    ws.append([])
    ws.append(["PII fields tracked", c.pii_field_count])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8


# ---------- Sheets ----------


_RULES_HEADERS = (
    "Rule UID",
    "Rule",
    "Severity",
    "Category",
    "Asset UID",
    "Asset",
    "Enforcement",
    "Field path",
)


def _write_rules(ws: Worksheet, rows: list[RuleApplication]) -> None:
    _write_table(
        ws,
        headers=_RULES_HEADERS,
        rows=[
            (
                r.rule_uid,
                r.rule_name,
                r.severity,
                r.category,
                r.asset_uid,
                r.asset_name,
                r.enforcement,
                r.field_path,
            )
            for r in rows
        ],
    )


_PII_HEADERS = ("Asset UID", "Asset", "Field path", "Data type", "Description", "Owners")


def _write_pii(ws: Worksheet, rows: list[PIIField]) -> None:
    _write_table(
        ws,
        headers=_PII_HEADERS,
        rows=[
            (
                r.asset_uid,
                r.asset_name,
                r.field_path,
                r.data_type,
                r.description,
                r.owner_names,
            )
            for r in rows
        ],
    )


_OWNERSHIP_HEADERS = (
    "Actor UID",
    "Actor",
    "Actor type",
    "Asset UID",
    "Asset",
    "Asset type",
)


def _write_ownership(ws: Worksheet, rows: list[OwnershipRow]) -> None:
    _write_table(
        ws,
        headers=_OWNERSHIP_HEADERS,
        rows=[
            (r.actor_uid, r.actor_name, r.actor_type, r.asset_uid, r.asset_name, r.asset_type)
            for r in rows
        ],
    )


_VERIFICATIONS_HEADERS = (
    "Kind",
    "UID",
    "Name",
    "Type / severity",
    "Discovered by",
    "Last updated",
)


def _write_verifications(ws: Worksheet, rows: list[VerificationEntry]) -> None:
    _write_table(
        ws,
        headers=_VERIFICATIONS_HEADERS,
        rows=[
            (
                r.entity_kind,
                r.entity_uid,
                r.entity_name,
                r.entity_type,
                r.discovered_by,
                r.updated_at,
            )
            for r in rows
        ],
    )


# ---------- helpers ----------


def _write_table(
    ws: Worksheet,
    *,
    headers: tuple[str, ...],
    rows: list[tuple[str, ...]],
) -> None:
    """Write a header row + data rows + auto-sized columns."""
    ws.append(list(headers))
    _stripe_header(ws, row=1, cols=len(headers))
    for row in rows:
        ws.append(list(row))

    for col_idx in range(1, len(headers) + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 80)


def _stripe_header(ws: Worksheet, *, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
