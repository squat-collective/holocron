"""Build Excel fixture files programmatically — no binary files in the repo."""

from collections.abc import Iterator
from pathlib import Path

import openpyxl
import pytest
from openpyxl.packaging.custom import CustomPropertyList, StringProperty
from openpyxl.worksheet.table import Table, TableStyleInfo


@pytest.fixture
def simple_xlsx(tmp_path: Path) -> Path:
    """One sheet, one ListObject table, header + 3 rows, no formulas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sales"
    ws.append(["id", "name", "amount"])
    ws.append([1, "Widget", 9.99])
    ws.append([2, "Gadget", 19.50])
    ws.append([3, "Doohickey", 4.25])

    table = Table(displayName="SalesTable", ref="A1:C4")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2")
    ws.add_table(table)

    path = tmp_path / "simple.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def heuristic_xlsx(tmp_path: Path) -> Path:
    """One sheet with a clear header row but NO Excel ListObject — heuristic territory."""
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Inventory"
    ws.append(["sku", "qty", "price"])
    for i in range(1, 6):
        ws.append([f"SKU-{i}", i * 10, i * 1.5])

    path = tmp_path / "heuristic.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def multi_sheet_xlsx(tmp_path: Path) -> Path:
    """Two sheets where sheet B references sheet A via VLOOKUP — produces lineage."""
    wb = openpyxl.Workbook()
    src = wb.active
    assert src is not None
    src.title = "Customers"
    src.append(["id", "name"])
    src.append([1, "Alice"])
    src.append([2, "Bob"])
    src.append([3, "Charlie"])
    src.add_table(Table(displayName="CustomersTable", ref="A1:B4"))

    dst = wb.create_sheet("Orders")
    dst.append(["order_id", "customer_id", "customer_name"])
    dst.append([100, 1, "=VLOOKUP(B2, Customers!A:B, 2, FALSE)"])
    dst.append([101, 2, "=VLOOKUP(B3, Customers!A:B, 2, FALSE)"])
    dst.add_table(Table(displayName="OrdersTable", ref="A1:C3"))

    path = tmp_path / "multi_sheet.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def metadata_xlsx(tmp_path: Path) -> Path:
    """Workbook with rich metadata: real creator, manager, and a custom 'Owner' prop."""
    wb = openpyxl.Workbook()
    wb.properties.creator = "Jean Dupont"
    wb.properties.lastModifiedBy = "Marie Curie"
    wb.properties.title = "Q4 Sales Report"
    wb.properties.description = "Quarterly sales tracking"

    custom = CustomPropertyList()
    custom.append(StringProperty(name="Owner", value="finance.team@acme.com"))
    custom.append(StringProperty(name="Department", value="Finance"))
    wb.custom_doc_props = custom

    ws = wb.active
    assert ws is not None
    ws.title = "Data"
    ws.append(["x", "y"])
    ws.append([1, 2])

    path = tmp_path / "metadata.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def noisy_creator_xlsx(tmp_path: Path) -> Path:
    """Workbook whose creator is the generic 'Microsoft Office User' — should be filtered."""
    wb = openpyxl.Workbook()
    wb.properties.creator = "Microsoft Office User"
    wb.properties.lastModifiedBy = "Real Person"

    ws = wb.active
    assert ws is not None
    ws.append(["a"])
    ws.append([1])

    path = tmp_path / "noisy.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def autouse_unit_only(request: pytest.FixtureRequest) -> Iterator[None]:
    """Marker no-op so each test gets a fresh tmp_path scope without ceremony."""
    yield
