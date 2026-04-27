"""Table detection — Excel's official ListObjects + heuristic rectangle detection."""

from typing import TYPE_CHECKING

from openpyxl.utils import get_column_letter

from excel_connector.columns import infer_column
from excel_connector.formulas import parse_formula
from excel_connector.models import (
    DetectedFormula,
    DetectedTable,
    TableConfidence,
)

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


def detect_tables(ws: "Worksheet") -> list[DetectedTable]:
    """Find tables on a worksheet.

    Two passes:
      1. Excel ListObjects (`ws.tables`) — confidence: certain
      2. Heuristic: contiguous rectangle of non-empty cells with a header row,
         only kept if it doesn't overlap with an already-detected ListObject.
    """
    detected: list[DetectedTable] = []
    occupied_ranges: list[tuple[int, int, int, int]] = []  # (min_row, min_col, max_row, max_col)

    # Pass 1: ListObjects
    # ws.tables.items() yields (name, ref_string); the full Table object lives at ws.tables[name].
    for table_name in list(ws.tables):
        table = ws.tables[table_name]
        ref = getattr(table, "ref", None) or ws.tables.get(table_name)
        if not ref:
            continue
        bounds = _parse_range(ref)
        if bounds is None:
            continue
        occupied_ranges.append(bounds)

        detected.append(
            _build_table(
                ws,
                name=table_name,
                bounds=bounds,
                confidence=TableConfidence.CERTAIN,
            )
        )

    # Pass 2: heuristic — find a contiguous rectangle starting from row 1, only if no
    # ListObject covers (1,1). This catches the classic "header in row 1, data below" sheet.
    if not _bounds_overlap(occupied_ranges, (1, 1, 1, 1)):
        bounds = _detect_heuristic_table(ws)
        if bounds is not None:
            detected.append(
                _build_table(
                    ws,
                    name=f"{ws.title}!table_1",
                    bounds=bounds,
                    confidence=TableConfidence.INFERRED,
                )
            )

    return detected


def _parse_range(ref: str) -> tuple[int, int, int, int] | None:
    """Parse 'A1:F100' into (min_row, min_col, max_row, max_col). Returns None if unparseable."""
    from openpyxl.utils import range_boundaries

    try:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
    except (ValueError, TypeError):
        return None
    return (min_row or 1, min_col or 1, max_row or 1, max_col or 1)


def _bounds_overlap(existing: list[tuple[int, int, int, int]], probe: tuple[int, int, int, int]) -> bool:
    pr1, pc1, pr2, pc2 = probe
    for er1, ec1, er2, ec2 in existing:
        if pr1 <= er2 and pr2 >= er1 and pc1 <= ec2 and pc2 >= ec1:
            return True
    return False


def _detect_heuristic_table(ws: "Worksheet") -> tuple[int, int, int, int] | None:
    """Find the rectangular region anchored at A1 with non-empty contiguous cells.

    Strategy: take row 1 as a candidate header. Walk right while cells are non-empty;
    that's the column extent. Walk down while at least one of those header columns has
    a non-empty cell; that's the row extent.

    Returns None if row 1 has no header content or fewer than 2 rows of data.
    """
    if ws.max_row is None or ws.max_row < 2:
        return None

    # Find header extent on row 1
    max_col = 0
    for col in range(1, (ws.max_column or 0) + 1):
        if ws.cell(row=1, column=col).value not in (None, ""):
            max_col = col
        else:
            break  # stop at first empty cell

    if max_col < 1:
        return None

    # Find row extent: walk down until we hit a fully-empty row across the header columns
    max_row = 1
    for row in range(2, (ws.max_row or 1) + 1):
        if any(ws.cell(row=row, column=c).value not in (None, "") for c in range(1, max_col + 1)):
            max_row = row
        else:
            break

    if max_row < 2:
        return None

    return (1, 1, max_row, max_col)


def _build_table(
    ws: "Worksheet",
    name: str,
    bounds: tuple[int, int, int, int],
    confidence: TableConfidence,
) -> DetectedTable:
    """Build a DetectedTable from worksheet bounds: read header, infer columns, scan formulas."""
    min_row, min_col, max_row, max_col = bounds

    # Headers from the first row of the bounds
    header_cells = [ws.cell(row=min_row, column=c) for c in range(min_col, max_col + 1)]
    header_names = [
        str(cell.value) if cell.value is not None else f"col_{i}"
        for i, cell in enumerate(header_cells)
    ]

    # Per-column values (rows below header)
    columns_data: list[list[object]] = [[] for _ in range(max_col - min_col + 1)]
    formulas: list[DetectedFormula] = []
    for row in range(min_row + 1, max_row + 1):
        for col_offset, col in enumerate(range(min_col, max_col + 1)):
            cell = ws.cell(row=row, column=col)
            value = cell.value
            columns_data[col_offset].append(value)
            if isinstance(value, str) and value.startswith("="):
                formulas.append(
                    parse_formula(
                        cell_address=f"{get_column_letter(col)}{row}",
                        formula=value,
                        current_sheet=ws.title,
                    )
                )

    columns = [
        infer_column(name=header_names[i], index=i, values=values)
        for i, values in enumerate(columns_data)
    ]

    range_str = (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )
    return DetectedTable(
        name=name,
        sheet_name=ws.title,
        confidence=confidence,
        range=range_str,
        row_count=max_row - min_row,  # excluding header
        columns=columns,
        formulas=formulas,
    )
