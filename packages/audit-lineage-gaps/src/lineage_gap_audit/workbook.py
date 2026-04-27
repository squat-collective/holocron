"""Render an `AuditReport` as an xlsx workbook.

Layout — one tab per finding category, plus an Overview tab that
counts each. Sheets stay flat (no merged cells, no formulas) so the
output works as a CSV-friendly source of truth for downstream tools
like pivot tables or scripts."""

from __future__ import annotations

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from lineage_gap_audit.analysis import AuditReport, Finding

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F2937")  # slate-800

# Column order is shared across every findings sheet so the workbook
# reads consistently.
FINDING_COLUMNS = ("entity_kind", "entity_uid", "entity_name", "entity_type", "reason", "detail")
FINDING_HEADERS = ("Kind", "UID", "Name", "Type / severity", "Reason", "Detail")


def write_audit_xlsx(report: AuditReport, *, generated_at: datetime) -> bytes:
    """Build the workbook and return its bytes."""
    wb = Workbook()
    # Remove the default empty sheet — we'll add our own.
    default = wb.active
    if default is not None:
        wb.remove(default)

    _write_overview(wb.create_sheet("Overview"), report, generated_at)
    _write_findings(wb.create_sheet("Orphan assets"), report.orphan_assets)
    _write_findings(wb.create_sheet("Lineage dead-ends"), report.lineage_deadends)
    _write_findings(wb.create_sheet("Undocumented"), report.undocumented_assets)
    _write_findings(wb.create_sheet("Dangling rules"), report.dangling_rules)
    _write_findings(wb.create_sheet("Unverified"), report.unverified_entities)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_overview(
    ws: Worksheet, report: AuditReport, generated_at: datetime
) -> None:
    ws.append(["Holocron lineage gap audit"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Generated at {generated_at.isoformat()}"])
    ws.append([])
    ws.append(["Category", "Count"])
    for cell in ws[4]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.append(["Orphan assets", len(report.orphan_assets)])
    ws.append(["Lineage dead-ends", len(report.lineage_deadends)])
    ws.append(["Undocumented assets", len(report.undocumented_assets)])
    ws.append(["Dangling rules", len(report.dangling_rules)])
    ws.append(["Unverified entities", len(report.unverified_entities)])
    ws.append([])
    ws.append(["Total findings", report.total_findings])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).font = Font(bold=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12


def _write_findings(ws: Worksheet, findings: list[Finding]) -> None:
    """Write a flat findings sheet. Header row is bold + dark, data rows
    use auto-fitted column widths."""
    ws.append(list(FINDING_HEADERS))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")

    for finding in findings:
        ws.append([_field(finding, col) for col in FINDING_COLUMNS])

    # Auto-size columns based on the rendered content. openpyxl doesn't
    # offer a true autosize, so we eyeball the max-cell-length per column
    # and clamp it to a readable cap.
    for col_idx, _ in enumerate(FINDING_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(12, max_len + 2), 80)


def _field(finding: Finding, name: str) -> str:
    """Read a `Finding` field by name. Keeping this explicit (rather than
    `getattr`) means a typo in `FINDING_COLUMNS` becomes a missing-attr
    error at first call rather than a silent blank cell."""
    if name == "entity_kind":
        return finding.entity_kind
    if name == "entity_uid":
        return finding.entity_uid
    if name == "entity_name":
        return finding.entity_name
    if name == "entity_type":
        return finding.entity_type
    if name == "reason":
        return finding.reason
    if name == "detail":
        return finding.detail
    raise KeyError(f"Unknown finding column: {name}")
