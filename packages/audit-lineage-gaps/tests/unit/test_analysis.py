"""Unit tests for the audit analyzers + workbook smoke check."""

from __future__ import annotations

import io
import zipfile
from datetime import UTC, datetime

from lineage_gap_audit.analysis import (
    find_dangling_rules,
    find_lineage_deadends,
    find_orphan_assets,
    find_undocumented_assets,
    find_unverified_entities,
    run_audit,
)
from lineage_gap_audit.models import (
    ActorRecord,
    AssetRecord,
    CatalogSnapshot,
    RelationRecord,
    RuleRecord,
)
from lineage_gap_audit.workbook import write_audit_xlsx

NOW = datetime(2026, 4, 25, 12, 0, 0, tzinfo=UTC)


def _asset(uid: str, **over: object) -> AssetRecord:
    base = {
        "uid": uid,
        "type": "dataset",
        "name": uid,
        "status": "active",
        "verified": True,
        "created_at": NOW,
        "updated_at": NOW,
    }
    base.update(over)  # type: ignore[arg-type]
    return AssetRecord.model_validate(base)


def _actor(uid: str, **over: object) -> ActorRecord:
    base = {
        "uid": uid,
        "type": "person",
        "name": uid,
        "verified": True,
        "created_at": NOW,
        "updated_at": NOW,
    }
    base.update(over)  # type: ignore[arg-type]
    return ActorRecord.model_validate(base)


def _rule(uid: str, **over: object) -> RuleRecord:
    base = {
        "uid": uid,
        "name": uid,
        "severity": "warning",
        "verified": True,
        "created_at": NOW,
        "updated_at": NOW,
    }
    base.update(over)  # type: ignore[arg-type]
    return RuleRecord.model_validate(base)


def _rel(from_uid: str, to_uid: str, type_: str) -> RelationRecord:
    return RelationRecord.model_validate(
        {
            "uid": f"rel-{from_uid}-{type_}-{to_uid}",
            "from_uid": from_uid,
            "to_uid": to_uid,
            "type": type_,
            "verified": True,
            "created_at": NOW,
        }
    )


# ---------- orphans ----------


class TestFindOrphanAssets:
    def test_asset_with_owns_relation_is_not_orphan(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a1")],
            actors=[_actor("p1")],
            relations=[_rel("p1", "a1", "owns")],
        )
        assert find_orphan_assets(snap) == []

    def test_asset_with_no_owns_is_orphan(self) -> None:
        snap = CatalogSnapshot(fetched_at=NOW, assets=[_asset("a1")])
        result = find_orphan_assets(snap)
        assert len(result) == 1
        assert result[0].entity_uid == "a1"
        assert result[0].reason == "No owner"

    def test_only_owns_counts_as_ownership(self) -> None:
        # `feeds` doesn't make a1 owned — it's a lineage edge, not an
        # ownership claim. Easy to confuse if someone ever skims the code.
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a1")],
            relations=[_rel("a2", "a1", "feeds")],
        )
        assert len(find_orphan_assets(snap)) == 1


# ---------- lineage dead-ends ----------


class TestFindLineageDeadends:
    def test_asset_with_no_feeds_or_uses_is_deadend(self) -> None:
        snap = CatalogSnapshot(fetched_at=NOW, assets=[_asset("a")])
        assert len(find_lineage_deadends(snap)) == 1

    def test_incoming_feeds_breaks_deadend(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a"), _asset("b")],
            relations=[_rel("b", "a", "feeds")],
        )
        # b is still a deadend (no incoming, only outgoing... wait, b does
        # have an outgoing feeds. b is NOT a deadend.).
        # a has incoming feeds — a is NOT a deadend.
        assert find_lineage_deadends(snap) == []

    def test_owns_relation_does_not_break_deadend(self) -> None:
        # `owns` doesn't carry data — the asset is still lineage-orphaned
        # even if it has a clearly-defined owner.
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a")],
            actors=[_actor("p")],
            relations=[_rel("p", "a", "owns")],
        )
        assert len(find_lineage_deadends(snap)) == 1


# ---------- undocumented ----------


class TestFindUndocumentedAssets:
    def test_no_description(self) -> None:
        snap = CatalogSnapshot(fetched_at=NOW, assets=[_asset("a", description=None)])
        assert len(find_undocumented_assets(snap)) == 1

    def test_empty_description(self) -> None:
        snap = CatalogSnapshot(fetched_at=NOW, assets=[_asset("a", description="")])
        assert len(find_undocumented_assets(snap)) == 1

    def test_whitespace_only_description_counts_as_empty(self) -> None:
        snap = CatalogSnapshot(fetched_at=NOW, assets=[_asset("a", description="   \n\t")])
        assert len(find_undocumented_assets(snap)) == 1

    def test_real_description_is_documented(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW, assets=[_asset("a", description="Sales rollup")]
        )
        assert find_undocumented_assets(snap) == []


# ---------- dangling rules ----------


class TestFindDanglingRules:
    def test_rule_without_applies_to_is_dangling(self) -> None:
        snap = CatalogSnapshot(fetched_at=NOW, rules=[_rule("r1")])
        result = find_dangling_rules(snap)
        assert len(result) == 1
        assert result[0].entity_uid == "r1"
        assert result[0].reason.startswith("Not applied")

    def test_rule_with_applies_to_is_not_dangling(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a")],
            rules=[_rule("r1")],
            relations=[_rel("r1", "a", "applies_to")],
        )
        assert find_dangling_rules(snap) == []


# ---------- unverified roll-up ----------


class TestFindUnverifiedEntities:
    def test_collects_unverified_across_kinds(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a", verified=False, discovered_by="csv")],
            actors=[_actor("p", verified=False)],
            rules=[_rule("r", verified=False)],
        )
        findings = find_unverified_entities(snap)
        kinds = sorted(f.entity_kind for f in findings)
        assert kinds == ["actor", "asset", "rule"]
        # Discovered-by hint propagates so the reviewer knows where the
        # record originated.
        asset_finding = next(f for f in findings if f.entity_kind == "asset")
        assert "csv" in asset_finding.detail

    def test_skips_verified_entities(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a", verified=True)],
            actors=[_actor("p", verified=True)],
            rules=[_rule("r", verified=True)],
        )
        assert find_unverified_entities(snap) == []


# ---------- run_audit roll-up ----------


class TestRunAudit:
    def test_total_findings_matches_sum(self) -> None:
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a", verified=False)],  # orphan + deadend + unverified
            rules=[_rule("r")],  # dangling
        )
        report = run_audit(snap)
        # a appears in orphan_assets + lineage_deadends + undocumented +
        # unverified. r is dangling.
        assert report.total_findings == (
            len(report.orphan_assets)
            + len(report.lineage_deadends)
            + len(report.undocumented_assets)
            + len(report.dangling_rules)
            + len(report.unverified_entities)
        )


# ---------- workbook smoke ----------


class TestWriteAuditXlsx:
    def test_produces_a_valid_xlsx_archive(self) -> None:
        # xlsx is a zip — checking the central directory is the cheapest
        # way to confirm openpyxl wrote a non-corrupt file without parsing
        # all the XML.
        snap = CatalogSnapshot(
            fetched_at=NOW,
            assets=[_asset("a")],
            rules=[_rule("r")],
        )
        body = write_audit_xlsx(run_audit(snap), generated_at=NOW)
        assert body[:2] == b"PK"  # zip magic
        with zipfile.ZipFile(io.BytesIO(body)) as zf:
            names = zf.namelist()
        assert any(n.endswith("workbook.xml") for n in names)

    def test_includes_expected_sheets(self) -> None:
        body = write_audit_xlsx(
            run_audit(CatalogSnapshot(fetched_at=NOW)), generated_at=NOW
        )
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(body))
        assert wb.sheetnames == [
            "Overview",
            "Orphan assets",
            "Lineage dead-ends",
            "Undocumented",
            "Dangling rules",
            "Unverified",
        ]
